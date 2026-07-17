"""IA (Investment Association) fund list import and name-to-ISIN resolution.

The IA publishes fund lists with names, management companies and sectors but no
ISINs. This module reads such a workbook, resolves each row to a Morningstar
security via :func:`fundlens.data.resolver.search_funds`, scores the candidate
matches, picks a representative share class, and writes an auditable resolution
file that can feed the alpha screener.

Resolution is deliberately conservative: every row is written to the audit CSV
as ``matched``/``ambiguous``/``unresolved`` so nothing is silently dropped, and
ambiguous/unresolved rows are also written to a review CSV for manual triage.
"""
from __future__ import annotations

import csv
import json
import re
import time
from collections.abc import Callable
from dataclasses import asdict, dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Literal

import openpyxl

from fundlens.cache import DiskCache
from fundlens.config import get_settings
from fundlens.data.resolver import FundSearchResult, search_funds

# ---------------------------------------------------------------------------
# Tunable constants
# ---------------------------------------------------------------------------

MATCHED_THRESHOLD = 0.86
AMBIGUOUS_THRESHOLD = 0.72
_SHARE_CLASS_TIE_DELTA = 0.02  # candidates within this of the top score count as one fund
_CACHE_TTL_DAYS = 30
_CHECKPOINT_INTERVAL = 25  # write checkpoint every N resolutions
_DEFAULT_LIMIT = 10
_TOP_CANDIDATES_FOR_AUDIT = 5

# Search fields required in the workbook header.
_HEADER_FUND_NAME = "Fund Name"
_HEADER_COMPANY = "Fund Management Company"
_HEADER_SECTOR = "Sector"

# Statuses that should be manually reviewed.
REVIEW_STATUSES = ("ambiguous", "unresolved")

# Audit CSV column order.
AUDIT_FIELDNAMES = [
    "ia_row_number",
    "ia_fund_name",
    "ia_management_company",
    "ia_sector",
    "status",
    "confidence",
    "isin",
    "resolved_name",
    "resolved_currency",
    "resolved_category",
    "security_type",
    "n_candidates",
    "top_candidates_json",
    "error",
]


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class IAFundRow:
    """One fund row imported from the IA workbook."""

    row_number: int  # original Excel row (data starts at row 2)
    fund_name: str
    management_company: str | None
    sector: str | None


@dataclass
class IAResolution:
    """The result of resolving one :class:`IAFundRow` to a Morningstar security."""

    ia_row_number: int
    ia_fund_name: str
    ia_management_company: str | None
    ia_sector: str | None
    status: Literal["matched", "ambiguous", "unresolved"]
    confidence: float | None
    isin: str | None
    resolved_name: str | None
    resolved_currency: str | None
    resolved_category: str | None  # not available from search; filled at screen time
    security_type: str | None
    candidates: list[dict] = field(default_factory=list)
    error: str | None = None


# ---------------------------------------------------------------------------
# Name normalisation
# ---------------------------------------------------------------------------

# Free-standing currency codes that should be removed from a name. Geographic
# tokens (asia, uk, global, ...) are deliberately NOT removed because they are
# the discriminative content for matching.
_CURRENCY_TOKENS = {
    "gbp", "gbx", "eur", "usd", "chf", "jpy", "aud", "cad", "hkd", "sgd",
    "sek", "nok", "dkk", "nzd", "cny", "inr", "krw", "twd", "zar", "pln",
    "czk", "huf", "ils", "mxn", "brl", "rub", "try",
}

# Structural wrappers that add no discriminative signal.
_WRAPPER_TOKENS = {
    "fund", "funds", "oeic", "sicav", "icvc", "plc", "portfolio",
    "trust", "ucits", "acc", "inc", "accumulation", "income",
    "class", "series",
}

# Bracketed content that is purely share-class / currency / domicile noise.
# This does NOT remove bracketed geography such as "(ex-Japan)" because the
# pattern requires the bracket to consist only of the listed tokens.
_BRACKET_NOISE = re.compile(
    r"\s*\((?:acc|accumulation|inc|income|dis|distribution|cap|a|b|c|i|z|r|t|"
    r"gbp|gbx|eur|usd|chf|jpy|hedged|unhedged|lux|ie|uk|usd acc|gbp acc|"
    r"eur acc|gbp inc|eur inc|usd inc|clean|institutional)\)",
    flags=re.IGNORECASE,
)

# Trailing share-class suffixes such as "- A (ACC)", "- Class A GBP", "- A".
# Anchored to end-of-string so they only strip trailing noise.
_TRAILING_SHARE_CLASS = re.compile(
    r"\s*[-\u2013]\s*(?:"
    r"class\s+[a-z]\s*(?:\s+(?:gbp|eur|usd|chf|acc|inc|hedge(?:d)?)\s*)?"
    r"|[a-z]\s*(?:acc|accumulation|inc|income|hedge(?:d)?)?"
    r"|[a-z](?:\s+(?:acc|inc|hedge(?:d)?))?"
    r")\s*$",
    flags=re.IGNORECASE,
)


def _clean_share_class_noise(name: str) -> str:
    """Strip share-class / currency / domicile noise from a fund name.

    Trailing share-class suffixes (``- A``, ``- Class A GBP``) and bracketed
    noise (``(ACC)``, ``(LUX)``) often co-occur (``- A (ACC)``), so the two
    passes are alternated until the name stops changing.
    """
    for _ in range(4):
        new = _TRAILING_SHARE_CLASS.sub("", name)
        new = _BRACKET_NOISE.sub("", new).strip()
        if new == name:
            break
        name = new
    return name.strip()


def _normalize_name(name: str) -> str:
    """Normalise a fund name for fuzzy matching.

    The goal is to make Morningstar's share-class-level names comparable to the
    IA's fund-level names, which differ mainly by share-class letters, currency,
    and structural wrappers. Geographic content is preserved.
    """
    if not name:
        return ""
    text = name.lower()
    text = _clean_share_class_noise(text)
    # Tokenise on non-alphanumerics so "ex-japan" -> "ex japan".
    tokens = re.findall(r"[a-z0-9]+", text)
    kept = [
        tok
        for tok in tokens
        if tok not in _CURRENCY_TOKENS and tok not in _WRAPPER_TOKENS
    ]
    return " ".join(kept)


def _normalize_token_set(name: str) -> set[str]:
    """Return the set of distinctive normalised tokens for a name."""
    return set(_normalize_name(name).split())


def _short_manager(company: str | None) -> str:
    """First lowercased alphabetic word of the management company."""
    if not company:
        return ""
    match = re.search(r"[A-Za-z]+", company)
    return match.group(0).lower() if match else ""


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------


def _score(
    ia_row: IAFundRow,
    candidate: FundSearchResult,
    *,
    manager_hint: str = "",
) -> float:
    """Score the similarity of a candidate to an IA row in ``[0, 1]``.

    Blends whole-string similarity (``difflib``) with distinctive token overlap,
    then applies a small boost if the manager's leading token appears in the
    candidate name.
    """
    ia_norm = _normalize_name(ia_row.fund_name)
    cand_norm = _normalize_name(candidate.name)
    if not ia_norm or not cand_norm:
        return 0.0

    base = SequenceMatcher(None, ia_norm, cand_norm).ratio()

    ia_tokens = set(ia_norm.split())
    cand_tokens = set(cand_norm.split())
    if ia_tokens:
        overlap = len(ia_tokens & cand_tokens) / len(ia_tokens)
    else:
        overlap = 0.0

    score = 0.6 * base + 0.4 * overlap

    if manager_hint and manager_hint in cand_norm:
        score += 0.06

    return min(score, 1.0)


# ---------------------------------------------------------------------------
# Share-class selection
# ---------------------------------------------------------------------------

# Tokens that indicate an institutional / clean share class. Single letters
# (i, z, x) are common institutional designations in European fund families;
# they are only matched in the trailing tokens of the name (see
# ``_share_class_rank``) so a fund-family tag like "FCP I" mid-name does not
# falsely flag retail share classes as institutional.
_INSTITUTIONAL_TOKENS = {
    "institutional", "inst", "clean", "i", "z", "x", "r6", "iacc", "zacc", "xacc",
}
_ACCUMULATION_TOKENS = {"acc", "accumulation", "aacc"}
_SHARE_CLASS_TAIL = 3  # number of trailing tokens inspected for share-class letters


def _share_class_rank(candidate: FundSearchResult) -> tuple[int, int, int]:
    """Preference tuple for a candidate share class (lower is better, element-wise).

    Implements the policy "prefer institutional/clean; otherwise prefer GBP;
    otherwise prefer accumulation". Returning a tuple means all three
    dimensions apply simultaneously, so an institutional GBP accumulation share
    class beats an institutional EUR one, which beats a retail GBP one.
    """
    name_norm = _normalize_name(candidate.name)
    tail = set(name_norm.split()[-_SHARE_CLASS_TAIL:])
    is_institutional = 0 if (tail & _INSTITUTIONAL_TOKENS) else 1
    raw_tokens = set(re.findall(r"[a-z0-9]+", candidate.name.lower()))
    is_accumulation = 0 if (raw_tokens & _ACCUMULATION_TOKENS) else 1
    is_gbp = 0 if (candidate.currency or "").upper() == "GBP" else 1
    return (is_institutional, is_gbp, is_accumulation)


def _select_share_class(
    scored: list[tuple[float, FundSearchResult]],
) -> tuple[float, FundSearchResult] | None:
    """Pick the representative candidate from scored share classes.

    Candidates tied within ``_SHARE_CLASS_TIE_DELTA`` of the top score are
    treated as share classes of the same fund and ranked by policy preference.
    The single best-matching fund otherwise wins outright.
    """
    if not scored:
        return None
    top_score = max(score for score, _ in scored)
    tied = [pair for pair in scored if pair[0] >= top_score - _SHARE_CLASS_TIE_DELTA]
    best = min(
        tied,
        key=lambda pair: (_share_class_rank(pair[1]), -pair[0], pair[1].name or ""),
    )
    return best


# ---------------------------------------------------------------------------
# Per-row resolution
# ---------------------------------------------------------------------------


def _build_queries(ia_row: IAFundRow) -> list[str]:
    """Search queries to try, in descending priority."""
    queries = [ia_row.fund_name.strip()]
    manager_word = _short_manager(ia_row.management_company)
    if manager_word and manager_word not in ia_row.fund_name.lower():
        queries.append(f"{ia_row.fund_name} {manager_word}")
    cleaned = _clean_share_class_noise(ia_row.fund_name)
    if cleaned and cleaned.lower() != ia_row.fund_name.strip().lower():
        queries.append(cleaned)
    # Dedupe preserving order and drop blanks.
    seen: set[str] = set()
    out: list[str] = []
    for q in queries:
        q_norm = q.strip()
        if q_norm and q_norm.lower() not in seen:
            seen.add(q_norm.lower())
            out.append(q_norm)
    return out


def _candidate_to_audit(candidate: FundSearchResult, score: float) -> dict:
    return {
        "isin": candidate.isin,
        "name": candidate.name,
        "currency": candidate.currency,
        "security_type": candidate.security_type,
        "score": round(score, 4),
    }


def _classify(score: float | None, *, matched: float, ambiguous: float) -> str:
    if score is None:
        return "unresolved"
    if score >= matched:
        return "matched"
    if score >= ambiguous:
        return "ambiguous"
    return "unresolved"


def resolve_ia_row(
    ia_row: IAFundRow,
    *,
    limit: int = _DEFAULT_LIMIT,
    matched_threshold: float = MATCHED_THRESHOLD,
    ambiguous_threshold: float = AMBIGUOUS_THRESHOLD,
    search_fn: Callable[[str, int], list[FundSearchResult]] | None = None,
) -> IAResolution:
    """Resolve a single IA row to a Morningstar security.

    Tries progressively cleaned queries until one returns candidates, scores
    each, picks a representative share class, and classifies by threshold.

    ``search_fn`` defaults to the module-level :func:`search_funds` resolved at
    call time, so tests can monkeypatch ``ia_universe.search_funds``.
    """
    if search_fn is None:
        search_fn = search_funds
    manager_hint = _short_manager(ia_row.management_company)
    last_error: str | None = None
    seen_isins: set[str] = set()
    collected: list[FundSearchResult] = []

    for query in _build_queries(ia_row):
        try:
            results = search_fn(query, limit)
        except Exception as exc:  # noqa: BLE001 - keep trying alternative queries
            last_error = f"{type(exc).__name__}: {exc}"
            continue
        for result in results or []:
            if result.isin and result.isin not in seen_isins:
                seen_isins.add(result.isin)
                collected.append(result)
        if collected:
            break

    if not collected:
        return IAResolution(
            ia_row_number=ia_row.row_number,
            ia_fund_name=ia_row.fund_name,
            ia_management_company=ia_row.management_company,
            ia_sector=ia_row.sector,
            status="unresolved",
            confidence=None,
            isin=None,
            resolved_name=None,
            resolved_currency=None,
            resolved_category=None,
            security_type=None,
            candidates=[],
            error=last_error or "no candidates returned",
        )

    scored = [
        (_score(ia_row, cand, manager_hint=manager_hint), cand) for cand in collected
    ]
    scored.sort(key=lambda pair: pair[0], reverse=True)

    chosen = _select_share_class(scored)
    if chosen is None:
        chosen_score, chosen_cand = scored[0]
    else:
        chosen_score, chosen_cand = chosen

    status = _classify(
        chosen_score, matched=matched_threshold, ambiguous=ambiguous_threshold
    )
    audit_candidates = [
        _candidate_to_audit(cand, score) for score, cand in scored[:_TOP_CANDIDATES_FOR_AUDIT]
    ]

    return IAResolution(
        ia_row_number=ia_row.row_number,
        ia_fund_name=ia_row.fund_name,
        ia_management_company=ia_row.management_company,
        ia_sector=ia_row.sector,
        status=status,
        confidence=round(chosen_score, 4),
        isin=chosen_cand.isin,
        resolved_name=chosen_cand.name,
        resolved_currency=chosen_cand.currency,
        resolved_category=None,
        security_type=chosen_cand.security_type,
        candidates=audit_candidates,
        error=None,
    )


# ---------------------------------------------------------------------------
# Workbook reading
# ---------------------------------------------------------------------------


def _find_header_columns(header: list[Any]) -> dict[str, int]:
    """Map required header names to their zero-based column indices."""
    mapping: dict[str, int] = {}
    normalized_to_target = {
        _HEADER_FUND_NAME.lower(): "fund_name",
        _HEADER_COMPANY.lower(): "management_company",
        _HEADER_SECTOR.lower(): "sector",
    }
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        key = normalized_to_target.get(str(cell).strip().lower())
        if key and key not in mapping:
            mapping[key] = idx
    missing = [name for name in ("fund_name", "management_company", "sector") if name not in mapping]
    if missing:
        friendly = {
            "fund_name": _HEADER_FUND_NAME,
            "management_company": _HEADER_COMPANY,
            "sector": _HEADER_SECTOR,
        }
        missing_names = ", ".join(friendly[name] for name in missing)
        raise ValueError(
            f"IA workbook header missing required column(s): {missing_names}. "
            f"Found header: {header}"
        )
    return mapping


def read_ia_workbook(
    path: str | Path,
    *,
    sheet: str = "Cleaned Fund List",
) -> list[IAFundRow]:
    """Read an IA workbook into a list of :class:`IAFundRow`.

    Header columns are located by name (not position) so column-order drift is
    tolerated. Whitespace is stripped; rows with a blank fund name are dropped.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(
            f"IA workbook has no sheet {sheet!r}; available: {wb.sheetnames}"
        )
    ws = wb[sheet]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration as exc:  # pragma: no cover - empty workbook
        raise ValueError(f"IA workbook sheet {sheet!r} is empty") from exc

    cols = _find_header_columns(header)
    name_col = cols["fund_name"]
    company_col = cols["management_company"]
    sector_col = cols["sector"]

    out: list[IAFundRow] = []
    for offset, row in enumerate(rows_iter, start=2):  # data rows start at Excel row 2
        if row is None:
            continue
        def _cell(idx: int) -> str | None:
            value = row[idx] if idx < len(row) else None
            if value is None:
                return None
            text = str(value).strip()
            return text or None

        fund_name = _cell(name_col)
        if not fund_name:
            continue
        out.append(
            IAFundRow(
                row_number=offset,
                fund_name=fund_name,
                management_company=_cell(company_col),
                sector=_cell(sector_col),
            )
        )
    return out


# ---------------------------------------------------------------------------
# Batch driver with cache + checkpoint/resume
# ---------------------------------------------------------------------------


def _cache_key(ia_row: IAFundRow) -> str:
    name_bit = _normalize_name(ia_row.fund_name) or "blank"
    mgr_bit = _short_manager(ia_row.management_company) or "blank"
    return f"ia_resolution/{name_bit}|{mgr_bit}"


def _load_checkpoint(checkpoint_path: Path | None) -> dict[int, IAResolution]:
    if checkpoint_path is None or not checkpoint_path.exists():
        return {}
    try:
        payload = json.loads(checkpoint_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    out: dict[int, IAResolution] = {}
    for row_number_str, item in payload.items():
        try:
            out[int(row_number_str)] = IAResolution(**item)
        except (TypeError, ValueError):
            continue
    return out


def _write_checkpoint(
    checkpoint_path: Path, resolutions: dict[int, IAResolution]
) -> None:
    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {str(k): asdict(v) for k, v in resolutions.items()}
    checkpoint_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def resolve_ia_workbook(
    path: str | Path,
    *,
    matched_threshold: float = MATCHED_THRESHOLD,
    ambiguous_threshold: float = AMBIGUOUS_THRESHOLD,
    limit: int = _DEFAULT_LIMIT,
    max_rows: int = 0,
    delay_seconds: float = 0.0,
    sheet: str = "Cleaned Fund List",
    use_cache: bool = True,
    checkpoint_path: str | Path | None = None,
    search_fn: Callable[[str, int], list[FundSearchResult]] | None = None,
    on_progress: Callable[[int, int, IAResolution], Any] | None = None,
) -> list[IAResolution]:
    """Resolve every row in an IA workbook.

    Resolution results are cached per row (keyed by normalised name + manager)
    for ``_CACHE_TTL_DAYS`` days, so re-running is cheap. When ``checkpoint_path``
    is provided, completed resolutions are also written there periodically and
    reloaded on start, so an interrupted long run can resume.

    ``search_fn`` defaults to the module-level :func:`search_funds` resolved at
    call time, so tests can monkeypatch ``ia_universe.search_funds``.
    """
    if search_fn is None:
        search_fn = search_funds
    ia_rows = read_ia_workbook(path, sheet=sheet)
    if max_rows > 0:
        ia_rows = ia_rows[:max_rows]
    total = len(ia_rows)

    # Call get_settings() inline so tests can monkeypatch the module-level
    # binding (mirrors the pattern in fundlens.data.universe.discover_fund_universe).
    cache = DiskCache(get_settings().cache_dir) if use_cache else None
    checkpoint = Path(checkpoint_path) if checkpoint_path else None
    completed = _load_checkpoint(checkpoint)

    results: list[IAResolution] = []
    for index, ia_row in enumerate(ia_rows, start=1):
        fresh = False  # True only when this row actually called Morningstar

        if ia_row.row_number in completed:
            resolution = completed[ia_row.row_number]
        elif cache is not None:
            cached = cache.get_json(_cache_key(ia_row), _CACHE_TTL_DAYS)
            if cached is not None:
                try:
                    resolution = IAResolution(**cached)
                except (TypeError, ValueError):
                    resolution = resolve_ia_row(
                        ia_row,
                        limit=limit,
                        matched_threshold=matched_threshold,
                        ambiguous_threshold=ambiguous_threshold,
                        search_fn=search_fn,
                    )
                    fresh = True
            else:
                resolution = resolve_ia_row(
                    ia_row,
                    limit=limit,
                    matched_threshold=matched_threshold,
                    ambiguous_threshold=ambiguous_threshold,
                    search_fn=search_fn,
                )
                fresh = True
                cache.put_json(_cache_key(ia_row), asdict(resolution))
        else:
            resolution = resolve_ia_row(
                ia_row,
                limit=limit,
                matched_threshold=matched_threshold,
                ambiguous_threshold=ambiguous_threshold,
                search_fn=search_fn,
            )
            fresh = True

        results.append(resolution)
        completed[ia_row.row_number] = resolution

        if checkpoint is not None and index % _CHECKPOINT_INTERVAL == 0:
            _write_checkpoint(checkpoint, completed)

        if on_progress is not None:
            on_progress(index, total, resolution)

        # Throttle only after live calls, so cached/checkpoint resumes are fast.
        if delay_seconds and fresh:
            time.sleep(delay_seconds)

    if checkpoint is not None:
        _write_checkpoint(checkpoint, completed)

    return results


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------


def _audit_row(resolution: IAResolution) -> dict:
    return {
        "ia_row_number": resolution.ia_row_number,
        "ia_fund_name": resolution.ia_fund_name,
        "ia_management_company": resolution.ia_management_company,
        "ia_sector": resolution.ia_sector,
        "status": resolution.status,
        "confidence": resolution.confidence,
        "isin": resolution.isin,
        "resolved_name": resolution.resolved_name,
        "resolved_currency": resolution.resolved_currency,
        "resolved_category": resolution.resolved_category,
        "security_type": resolution.security_type,
        "n_candidates": len(resolution.candidates),
        "top_candidates_json": json.dumps(resolution.candidates, ensure_ascii=False),
        "error": resolution.error,
    }


def write_resolution_audit(
    resolutions: list[IAResolution], out_path: str | Path
) -> Path:
    """Write every resolution (all statuses) to an audit CSV."""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=AUDIT_FIELDNAMES)
        writer.writeheader()
        for resolution in resolutions:
            writer.writerow(_audit_row(resolution))
    return out


def write_resolution_review(
    resolutions: list[IAResolution], out_path: str | Path
) -> Path:
    """Write ambiguous + unresolved resolutions to a review CSV."""
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    review = [r for r in resolutions if r.status in REVIEW_STATUSES]
    with open(out, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=AUDIT_FIELDNAMES)
        writer.writeheader()
        for resolution in review:
            writer.writerow(_audit_row(resolution))
    return out


def read_resolution_audit(path: str | Path) -> list[dict]:
    """Read an audit CSV back into row dicts (for screen-resolved)."""
    with open(path, "r", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def resolution_counts(resolutions: list[IAResolution]) -> dict[str, int]:
    """Return a {status: count} summary of resolutions."""
    counts: dict[str, int] = {"matched": 0, "ambiguous": 0, "unresolved": 0}
    for resolution in resolutions:
        counts[resolution.status] = counts.get(resolution.status, 0) + 1
    return counts


# ---------------------------------------------------------------------------
# Screen checkpointing (incremental per-fund persistence with resume)
# ---------------------------------------------------------------------------


def load_screen_checkpoint(
    path: str | Path, fieldnames: list[str]
) -> tuple[list[dict], set[str]]:
    """Load existing screen rows from a checkpoint CSV.

    Returns the completed rows (in file order) and the set of ISINs already
    done, so a resumed run can skip them. A missing or unreadable checkpoint
    yields an empty result, so callers can always (re)start cleanly.
    """
    p = Path(path)
    if not p.exists():
        return [], set()
    rows: list[dict] = []
    done: set[str] = set()
    try:
        with p.open("r", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                rows.append(row)
                isin = (row.get("isin") or "").strip()
                if isin:
                    done.add(isin)
    except OSError:
        return [], set()
    return rows, done


def append_screen_row(path: str | Path, fieldnames: list[str], row: dict) -> None:
    """Append one screen row to a checkpoint CSV, creating it (with header) if new.

    Each call opens, writes, and closes the file so the checkpoint is durable
    immediately after every fund, not just at the end of a run. An interrupted
    long screen therefore loses at most the in-flight row.
    """
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    write_header = not p.exists() or p.stat().st_size == 0
    with p.open("a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()
        writer.writerow({field: row.get(field) for field in fieldnames})
